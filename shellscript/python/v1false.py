# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl as opx
import os
import sys
import shutil
import xlrd
def is_num(s):
        return s.replace(',', '').replace('.', '').replace('-', '').isnumeric()



InputCaseName = sys.argv
FileDir = os.environ['HOME'] + "/ftpdoc"
ExcelPath =  FileDir + "/V1falseZyymmdd.xlsx"
NewExcelPath = FileDir + "/V1false" + InputCaseName[1]  + ".xlsx"
shutil.copyfile(ExcelPath,NewExcelPath)

# # 対象ファイルのExcelWriterの呼び出し
# EXL=pd.ExcelWriter(NewExcelPath, engine='openpyxl')
# # ExcelWriterに既存の対象ファイルを読み込ませる
# EXL.book=opx.load_workbook(NewExcelPath)
# # 既存のsheet情報を読み込ませる
# EXL.sheets=dict((ws.title, ws) for ws in EXL.book.worksheets )
#
#
# UTfileDir= FileDir +"/" + InputCaseName[1] + "/TCAZDOC/numberOfUntreatedTransaction.csv"
# UntreatedTransaction = pd.read_csv(UTfileDir,header=None)
# print ("read VerifyResult件数")
# # ExcelWriterを用いて新規シートにDataFrameを保存
# UntreatedTransaction.to_excel(EXL,'VerifyResult件数' ,index=False ,header=False)
# print("write VerifyResult件数")
#
#
#
#
# UTSfileDir= FileDir + "/"+ InputCaseName[1] + "/TCAZDOC/sumOfUntreatedTransactionSequenceNumber.csv"
# UntreatedTransactionSum = pd.read_csv(UTSfileDir,header=None)
# print("read VerifyResult シーケンス")
# # ExcelWriterを用いて新規シートにDataFrameを保存
# UntreatedTransactionSum.to_excel(EXL,'VerifyResult シーケンス',index=False ,header=False)
# print ("write VerifyResult件数")
#
#
#
# LOSTFNDfileDir= FileDir + "/"+ InputCaseName[1] + "/TCAZDOC/" + InputCaseName[1] +".LOSTFNDR.TXT.csv"
# LOSTFND = pd.read_csv(LOSTFNDfileDir,header=None)
# print("read LOSTFNDR")
# # ExcelWriterを用いて新規シートにDataFrameを保存
# LOSTFND.to_excel(EXL,'LOSTFNDR' ,index=False ,header=False)
# print ("write LOSTFNDR")
#
# EXL.save()
# EXL.close()
# ExcelWriterに既存の対象ファイルを読み込ませる

# wb = xlrd.open_workbook(NewExcelPath)
# sheet = wb.sheet_by_name('突き合わせ')
# print(sheet)
print(NewExcelPath)
workbook = opx.load_workbook(NewExcelPath, data_only=True)
sheet_names = workbook.sheetnames
sheet = workbook[sheet_names[0]]
print(sheet['B1'].value)
